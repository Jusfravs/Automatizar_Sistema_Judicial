# src/gestor_casos.py
import os
import sys
import json
import shutil
from datetime import datetime
import pandas as pd
from pandas.errors import EmptyDataError
from src.logger_config import obtener_logger

logger = obtener_logger("GestorCasos")

class GestorCasos:
    """
    Repositorio CRUD de datos para la lectura, actualización y persistencia del reporte.
    """
    COLUMNAS_MOLDE_EXPORTACION = [
        'FECHA INICIO JUICIO',
        'FECHA FIN ULTIMA FASE',
        'ULTIMA ETAPA',
        'ULTIMA FASE',
        'FECHA INICIO FASE ACTUAL',
        'ETAPA ACTUAL',
        'FASE ACTUAL',
        'DIAS TRANSCURRIDOS',
    ]
    COLUMNAS_FECHA_PROCESAL = {
        'FECHA FIN ULTIMA FASE',
        'FECHA INICIO FASE ACTUAL',
    }
    COLUMNAS_FASE_EXPORTACION = ('ULTIMA FASE', 'FASE ACTUAL')
    ETIQUETAS_FASE_EXCEL = {
        '2.1 CITACION (PERSONA/BOLETA)': '2.1 CITACION',
        '6.5 CONGELAMIENTO DE CUENTAS / CIERRE': (
            '6.5 CONGELAMIENTO DE CUENTAS'
        ),
    }

    @staticmethod
    def _parsear_fecha_reporte(valor):
        """Interpreta fechas del portal, incluidas marcas ISO con zona horaria."""
        if valor is None or pd.isna(valor) or str(valor).strip() == "":
            return None
        if isinstance(valor, pd.Timestamp):
            return valor.to_pydatetime()
        if isinstance(valor, datetime):
            return valor

        fecha_str = str(valor).strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(fecha_str[:10], fmt)
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(fecha_str.replace("Z", "+00:00"))
        except ValueError:
            return None

    @classmethod
    def _normalizar_fecha_reporte(cls, valor):
        fecha = cls._parsear_fecha_reporte(valor)
        return fecha.strftime("%d/%m/%Y") if fecha else valor

    @classmethod
    def _normalizar_fase_exportacion(cls, valor):
        """Acorta etiquetas únicamente para la presentación del Excel."""
        if valor is None or pd.isna(valor):
            return valor
        return cls.ETIQUETAS_FASE_EXCEL.get(str(valor).strip(), valor)

    def __init__(self, ruta_config="config.json"):
        with open(ruta_config, 'r', encoding='utf-8') as f:
            self.config = json.load(f)

        rutas = self.config.get('rutas', {})
        self.ruta_csv = rutas.get('archivo_csv', 'data/reporte_trabajo.csv')
        self.ruta_excel = rutas.get('archivo_origen', 'data/REPORTE JUICIOS PARA REVISIÓN JULIO.xlsx')
        self.ruta_final = rutas.get('archivo_excel_final', 'data/REPORTE_PROCESADO_FINAL.xlsx')
        self.hoja = rutas.get('hoja_lectura', 'migrado')
        self.filtros = self.config.get('filtros_activos', {})

        # Cargar CSV existente si existe y es válido
        if os.path.exists(self.ruta_csv):
            try:
                self.df = pd.read_csv(self.ruta_csv, low_memory=False)
                self.df.columns = self.df.columns.astype(str).str.strip().str.upper()
                if self.df.empty:
                    raise EmptyDataError("CSV vacío")
            except Exception as e:
                logger.warning("No se pudo cargar el CSV existente ('%s'): %s. Regenerando desde Excel...", self.ruta_csv, e)
                self._inicializar_csv(forzar=True)
                self.df = pd.read_csv(self.ruta_csv, low_memory=False)
                self.df.columns = self.df.columns.astype(str).str.strip().str.upper()
        else:
            logger.info("CSV no encontrado. Creando desde Excel original...")
            self._inicializar_csv(forzar=True)
            self.df = pd.read_csv(self.ruta_csv, low_memory=False)
            self.df.columns = self.df.columns.astype(str).str.strip().str.upper()

        # Comparar contra el total del reporte; otros archivos validos pueden
        # contener menos de 1000 filas.
        total_esperado = self.config.get('auditoria', {}).get('total_esperado')
        csv_incompleto = (
            total_esperado is not None and len(self.df) < int(total_esperado)
        )
        if ('SUCURSAL' not in self.df.columns or csv_incompleto) and os.path.exists(self.ruta_excel):
            logger.info("El CSV tiene %s registros. Sincronizando datos con el Excel completo...", len(self.df))
            self._inicializar_csv(forzar=True)
            self.df = pd.read_csv(self.ruta_csv, low_memory=False)
            self.df.columns = self.df.columns.astype(str).str.strip().str.upper()

    def _cargar_excel_robusto(self):
        """Carga el Excel usando una copia sombra para evitar bloqueos si está abierto en Excel, e infiere el header."""
        import subprocess
        excel_path = os.path.abspath(self.ruta_excel)
        temp_path = os.path.abspath(os.path.join(os.path.dirname(self.ruta_excel), "_temp_excel_shadow.xlsx"))
        
        # Copiar con PowerShell para evitar error de bloqueo de archivo exclusivo en Windows
        cmd = f'powershell -Command "Copy-Item \'{excel_path}\' \'{temp_path}\' -Force"'
        subprocess.run(cmd, shell=True, capture_output=True)

        archivo_lectura = temp_path if os.path.exists(temp_path) else self.ruta_excel

        try:
            for h in [0, 1]:
                try:
                    df = pd.read_excel(archivo_lectura, sheet_name=self.hoja, header=h)
                    df.columns = df.columns.astype(str).str.strip().str.upper()
                    if 'SUCURSAL' in df.columns or 'CODIGO_JUICIO' in df.columns:
                        logger.info("Excel cargado correctamente detectando header=%s (%s filas).", h, len(df))
                        return df
                except Exception:
                    continue
            raise ValueError("No se pudo detectar la cabecera correcta en el archivo Excel.")
        finally:
            if os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except Exception:
                    pass

    def _inicializar_csv(self, forzar=False):
        """CREATE: Genera o combina el CSV de trabajo desde el Excel original."""
        if not os.path.exists(self.ruta_excel):
            return

        if os.path.exists(self.ruta_csv) and not forzar:
            return

        logger.info("Inicializando/sincronizando base de datos CSV desde Excel...")
        df_excel = self._cargar_excel_robusto()

        if os.path.exists(self.ruta_csv):
            try:
                df_existente = pd.read_csv(self.ruta_csv, low_memory=False)
                df_existente.columns = df_existente.columns.astype(str).str.strip().str.upper()
                if 'CODIGO_JUICIO' in df_existente.columns and 'CODIGO_JUICIO' in df_excel.columns:
                    logger.info("Combinando datos existentes del CSV con la base completa del Excel...")
                    df_merged = df_existente.set_index('CODIGO_JUICIO').combine_first(df_excel.set_index('CODIGO_JUICIO')).reset_index()
                    df_merged.to_csv(self.ruta_csv, index=False, encoding='utf-8-sig')
                    return
            except Exception as e:
                logger.warning("No se pudo combinar el CSV existente, se creará uno nuevo: %s", e)

        df_excel.to_csv(self.ruta_csv, index=False, encoding='utf-8-sig')

    def obtener_casos_pendientes(self):
        """READ: Obtiene la lista de números de juicio que cumplen con los filtros."""
        logger.debug("Columnas disponibles: %s", self.df.columns.tolist())
        columna_configurada = str(self.filtros.get('columna_estado_judicial', '')).strip().upper()
        if columna_configurada:
            if columna_configurada not in self.df.columns:
                raise KeyError(
                    f"La columna configurada para estado judicial '{columna_configurada}' no existe en el CSV."
                )
            col_estado = columna_configurada
            logger.info("Usando columna configurada para estado judicial: '%s'.", col_estado)
        elif 'ESTADO' in self.df.columns:
            col_estado = 'ESTADO'
        elif 'ESTADO.1' in self.df.columns:
            col_estado = 'ESTADO.1'
            logger.warning("Se usará la columna de respaldo 'ESTADO.1'.")
        else:
            raise KeyError("No se encontro una columna 'ESTADO' ni 'ESTADO.1' en el CSV.")
        
        mask = pd.Series(True, index=self.df.index)

        suc = str(self.filtros.get('sucursal', '') or '').strip().upper()
        if suc and suc not in ('TODAS', 'TODOS', 'ALL', 'NONE'):
            mask &= (self.df['SUCURSAL'].astype(str).str.strip().str.upper() == suc)

        ofi = str(self.filtros.get('oficina', '') or '').strip().upper()
        if ofi and ofi not in ('TODAS', 'TODOS', 'ALL', 'NONE'):
            mask &= (self.df['OFICINA'].astype(str).str.strip().str.upper() == ofi)
        usuario = str(self.filtros.get('usuario', '') or '').strip().upper()

        if usuario and usuario not in ('TODAS', 'TODOS', 'ALL', 'NONE'):
            if 'USUARIO' not in self.df.columns:
                raise KeyError("La columna configurada para usuario 'USUARIO' no existe en el CSV.")
            mask &= (
                self.df['USUARIO'].astype(str).str.strip().str.upper() == usuario
            )

        est = str(self.filtros.get('estado_judicial', '') or '').strip().upper()
        if est and est not in ('TODAS', 'TODOS', 'ALL', 'NONE'):
            mask &= (self.df[col_estado].astype(str).str.strip().str.upper() == est)

        df_filtrado = self.df[mask]

        casos = df_filtrado['NUMERO_JUICIO'].dropna().astype(str).str.strip().tolist()

        # Aplicar punto de partida si fue especificado
        inicio = self.filtros.get('inicio_desde_juicio')
        if inicio:
            inicio_limpio = str(inicio).replace("-", "").strip()
            idx = next((i for i, c in enumerate(casos) if str(c).replace("-", "").strip() == inicio_limpio), None)
            if idx is not None:
                logger.info("Reanudando desde causa '%s' (Caso #%s de %s).", inicio, idx + 1, len(casos))
                casos = casos[idx:]

        return casos

    def actualizar_caso(self, numero_juicio, datos):
        """UPDATE: Inyecta en la fila correspondiente los datos extraídos."""
        numero_juicio_normalizado = str(numero_juicio).strip().upper()
        mask = (
            self.df['NUMERO_JUICIO'].astype(str).str.strip().str.upper()
            == numero_juicio_normalizado
        )
        if mask.any():
            for col, val in datos.items():
                if val is not None:
                    if col in self.COLUMNAS_FECHA_PROCESAL:
                        val = self._normalizar_fecha_reporte(val)
                    if col not in self.df.columns:
                        self.df[col] = None

                    # Si el valor es una lista o diccionario (ej. HISTORIAL_ACTUACIONES), serializar a JSON string
                    if isinstance(val, (list, dict)):
                        val = json.dumps(val, ensure_ascii=False)
                    # Un CSV con una columna completamente vacia se carga como
                    # float64. Pandas recientes rechazan guardar texto (por
                    # ejemplo, una fecha dd/mm/aaaa) sin convertir antes el tipo.
                    if isinstance(val, str) and self.df[col].dtype != object:
                        self.df[col] = self.df[col].astype(object)


                    self.df.loc[mask, col] = val
            return True
        return False

    def guardar(self):
        """SAVE: Persiste los cambios en la base CSV de trabajo. Retorna True en éxito."""
        if os.path.isfile(self.ruta_csv):
            ruta_backup = f"{self.ruta_csv}.bak"
            try:
                shutil.copy2(self.ruta_csv, ruta_backup)
                logger.debug("Respaldo del CSV creado en: %s", ruta_backup)
            except OSError as e:
                logger.error("No se pudo crear respaldo del CSV: %s", e)
                return False

        try:
            self.df.to_csv(self.ruta_csv, index=False, encoding='utf-8-sig')
            return True
        except Exception as e:
            logger.error("No se pudo guardar CSV: %s", e)
            return False

    def exportar_excel(self):
        """EXPORT: Genera el Excel .xlsx consolidado final con reestructuración de columnas y formato en rojo."""
        import openpyxl
        from openpyxl.styles import PatternFill, Font

        # 1. Calcular días transcurridos
        self.calcular_dias_transcurridos()

        # 2. Lista de las columnas obsoletas a eliminar del Excel final si existen
        cols_a_eliminar = [
            'ETAPA_PROCESAL (ACTUAL)',
            'FASE_PROCESAL (ACTUAL)',
            'ETAPA_PROCESAL (MIGRADO)',
            'CODIGO_FASE',
            'FASE_PROCESAL (MIGRADO)',
            'FECHA INICIAL FASE ACTUAL',
            'DIAS EN LA FASE ACTUAL',
            'ETAPA_PROCESAL',
            'FASE_PROCESAL'
        ]

        df_export = self.df.copy()

        # Crear copia de FECHA INICIO JUICIO si no existe
        if 'FECHA INICIO JUICIO' not in df_export.columns:
            df_export['FECHA INICIO JUICIO'] = None

        # Asegurar presencia de nuevas columnas MOLDE
        nuevas_cols_molde = list(self.COLUMNAS_MOLDE_EXPORTACION)

        for col in nuevas_cols_molde:
            if col not in df_export.columns:
                df_export[col] = None

        # Presentar nombres breves sin modificar las etiquetas procesales internas.
        for col_fase in self.COLUMNAS_FASE_EXPORTACION:
            df_export[col_fase] = df_export[col_fase].map(
                self._normalizar_fase_exportacion
            )

        # Normalizar timestamps ISO ya existentes sin volver a consultar el portal.
        for col_fecha in self.COLUMNAS_FECHA_PROCESAL:
            df_export[col_fecha] = df_export[col_fecha].map(
                self._normalizar_fecha_reporte
            )

        # Copiar FECHA FIN ULTIMA FASE a FECHA INICIO FASE ACTUAL si está vacía
        mask_copia = df_export['FECHA INICIO FASE ACTUAL'].isna() & df_export['FECHA FIN ULTIMA FASE'].notna()
        df_export.loc[mask_copia, 'FECHA INICIO FASE ACTUAL'] = df_export.loc[mask_copia, 'FECHA FIN ULTIMA FASE']

        # Eliminar columnas viejas de df_export
        cols_existentes_eliminar = [c for c in cols_a_eliminar if c in df_export.columns]
        df_export.drop(columns=cols_existentes_eliminar, inplace=True, errors='ignore')

        # Reordenar columnas: poner las 8 nuevas después de COMENTARIO_ULTIMO con 1 columna vacía separadora
        cols_base = [c for c in df_export.columns if c not in nuevas_cols_molde and c != ' ']

        if 'COMENTARIO_ULTIMO' in cols_base:
            idx_comentario = cols_base.index('COMENTARIO_ULTIMO')
            cols_izq = cols_base[:idx_comentario + 1]
            cols_der = cols_base[idx_comentario + 1:]
        else:
            cols_izq = cols_base
            cols_der = []

        df_export[' '] = ""  # Columna separadora vacía
        cols_ordenadas = cols_izq + cols_der + [' '] + nuevas_cols_molde
        
        # Eliminar posibles duplicados manteniendo orden
        cols_finales = []
        vistos = set()
        for c in cols_ordenadas:
            if c in df_export.columns and c not in vistos:
                cols_finales.append(c)
                vistos.add(c)

        df_final = df_export[cols_finales]

        logger.info("Exportando informe final reestructurado a: %s", self.ruta_final)
        try:
            df_final.to_excel(self.ruta_final, index=False, sheet_name=self.hoja)
        except PermissionError:
            from datetime import datetime as _dt
            base, ext = os.path.splitext(self.ruta_final)
            ruta_alt = f"{base}_{_dt.now().strftime('%Y%m%d_%H%M%S')}{ext}"
            logger.warning(
                "Archivo Excel bloqueado (%s). Guardando copia en: %s",
                self.ruta_final, ruta_alt,
            )
            df_final.to_excel(ruta_alt, index=False, sheet_name=self.hoja)
            self.ruta_final = ruta_alt

        # 3. Aplicar formato condicional a filas con error en rojo usando openpyxl
        try:
            wb = openpyxl.load_workbook(self.ruta_final)
            ws = wb[self.hoja] if self.hoja in wb.sheetnames else wb.active

            fill_rojo = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            font_rojo = Font(color="9C0006", bold=True)

            # Buscar índice de columna COMENTARIO_ULTIMO o cualquier celda con error
            col_comentario_idx = None
            for col_idx in range(1, ws.max_column + 1):
                header_val = ws.cell(row=1, column=col_idx).value
                if header_val and str(header_val).strip().upper() == 'COMENTARIO_ULTIMO':
                    col_comentario_idx = col_idx
                    break

            for row_idx in range(2, ws.max_row + 1):
                es_error = False
                if col_comentario_idx:
                    val = ws.cell(row=row_idx, column=col_comentario_idx).value
                    if val and ("ERROR:" in str(val).upper() or "NO DEVOLVIO RESULTADOS" in str(val).upper()):
                        es_error = True

                if es_error:
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.fill = fill_rojo
                        cell.font = font_rojo

            wb.save(self.ruta_final)
            wb.close()
            logger.info("Formato de resaltado rojo para errores aplicado correctamente.")
        except Exception as e_xl:
            logger.warning("No se pudo aplicar estilo openpyxl al Excel final: %s", e_xl)

        logger.info("¡Archivo Excel final generado exitosamente!")

    def calcular_dias_transcurridos(self, fecha_actual=None):
        """
        Calcula la columna 'DIAS TRANSCURRIDOS' como la diferencia en días calendario
        entre la fecha actual y 'FECHA FIN ULTIMA FASE' (o 'FECHA INICIAL FASE ACTUAL').
        Soporta formatos dd/mm/yyyy y yyyy-mm-dd.
        """
        col_fecha = 'FECHA FIN ULTIMA FASE' if 'FECHA FIN ULTIMA FASE' in self.df.columns else 'FECHA INICIAL FASE ACTUAL'
        col_dias = 'DIAS TRANSCURRIDOS'

        if col_fecha not in self.df.columns:
            logger.warning("Columna '%s' no encontrada. No se calculará '%s'.", col_fecha, col_dias)
            return

        self.df[col_dias] = None

        hoy = (fecha_actual or datetime.now()).date()
        conteo = 0

        for idx, valor in self.df[col_fecha].items():
            if pd.isna(valor) or str(valor).strip() == "":
                continue

            fecha_parsed = self._parsear_fecha_reporte(valor)

            if fecha_parsed:
                dias = (hoy - fecha_parsed.date()).days
                self.df.at[idx, col_dias] = max(0, dias)
                conteo += 1

        logger.info("'%s' calculado para %s registros.", col_dias, conteo)

    def calcular_dias_fase_actual(self):
        """Método de retrocompatibilidad que delega en calcular_dias_transcurridos."""
        self.calcular_dias_transcurridos()

